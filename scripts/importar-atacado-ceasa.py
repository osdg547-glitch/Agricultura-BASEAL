#!/usr/bin/env python3
"""Gera dados/precos-atacado-ceasa-se-2026.json a partir da planilha consolidada
da EMDAGRO/ASPLAN em dados/fontes/.

A planilha é a transcrição dos boletins "Preços médios de atacado - CEASA"
publicados em PDF. Este script não interpreta nem corrige preço nenhum: ele
apenas transpõe a aba "Série larga", converte para R$/kg quando a unidade de
origem tem peso nominal, e carrega junto a procedência de cada data (boletim de
origem e marca de realinhamento) lida da aba "Série longa".

Dependência: openpyxl (pip install openpyxl). Só é necessária para reimportar a
planilha; o site consome apenas o JSON gerado.

Uso:
    python3 scripts/importar-atacado-ceasa.py
"""

import json
import re
import unicodedata
from collections import OrderedDict
from pathlib import Path

import openpyxl

RAIZ = Path(__file__).resolve().parent.parent
PLANILHA = RAIZ / "dados" / "fontes" / "precos-atacado-ceasa-se-consolidado-v3.xlsx"
SAIDA = RAIZ / "dados" / "precos-atacado-ceasa-se-2026.json"

# Datas que aparecem no cabeçalho dos boletins marcadas como feriado, sem preço
# publicado. Não entram na série.
FERIADOS = ["2026-04-02", "2026-04-21", "2026-06-04"]

# Os sete últimos produtos do boletim formam o bloco de origem animal.
PRODUTOS_ANIMAIS = {
    "Carne bovina dianteira",
    "Carne bovina traseira",
    "Frango de granja",
    "Manteiga",
    "Ovo branco grande",
    "Ovo vermelho grande",
    "Queijo coalho",
}

# Slugs que já circulam no portal (fichas de produto e mural da capa) e não podem
# mudar, mais os casos em que a transliteração automática ficaria ilegível.
SLUGS_FIXOS = {
    "Macaxeira (Aipim)": "macaxeira",
    "Amendoim c/ casca cozido": "amendoim_com_casca_cozido",
}

# Rótulos já publicados no portal, mantidos para não divergir das fichas.
ROTULOS_FIXOS = {
    "Macaxeira (Aipim)": "Macaxeira (aipim)",
}

# Unidades de contagem: o boletim cota por peça, não por peso. Sem peso nominal
# publicado pela fonte, converter para R$/kg seria inventar número.
UNIDADES_CONTAGEM = {
    "cento": (100, "unidade"),
    "cx 30 dz": (360, "unidade"),
}


def slug(rotulo):
    if rotulo in SLUGS_FIXOS:
        return SLUGS_FIXOS[rotulo]
    sem_acento = unicodedata.normalize("NFKD", rotulo).encode("ascii", "ignore").decode()
    return re.sub(r"_+", "_", re.sub(r"[^a-z0-9]+", "_", sem_acento.lower())).strip("_")


def analisar_unidade(unidade):
    """Devolve (fator_kg, itens_por_unidade, tipo) para a unidade do boletim.

    fator_kg é None quando a unidade é de contagem: nesse caso o preço fica na
    unidade de origem e o portal precisa dizer isso ao leitor.
    """
    chave = unidade.strip().lower()
    if chave in UNIDADES_CONTAGEM:
        itens, tipo = UNIDADES_CONTAGEM[chave]
        return None, itens, tipo
    achado = re.search(r"(\d+(?:[.,]\d+)?)\s*kg", chave)
    if achado:
        return float(achado.group(1).replace(",", ".")), None, "peso"
    if chave == "kg":
        return 1.0, None, "peso"
    raise ValueError("unidade não reconhecida: %r" % unidade)


def arred(valor, casas=4):
    arredondado = round(valor, casas)
    return int(arredondado) if arredondado == int(arredondado) else arredondado


def medias_mensais(datas, precos):
    acumulado = OrderedDict()
    for data, preco in zip(datas, precos):
        if preco is None:
            continue
        acumulado.setdefault(data[:7], []).append(preco)
    return OrderedDict((mes, arred(sum(v) / len(v))) for mes, v in acumulado.items())


def estatisticas(precos):
    serie = [p for p in precos if p is not None]
    media = sum(serie) / len(serie)
    # Desvio-padrão amostral, como na aba de estatísticas da planilha de origem.
    desvio = (sum((p - media) ** 2 for p in serie) / (len(serie) - 1)) ** 0.5
    mudancas = sum(1 for a, b in zip(serie, serie[1:]) if a != b)
    return OrderedDict([
        ("n_observacoes", len(serie)),
        ("minimo", arred(min(serie))),
        ("maximo", arred(max(serie))),
        ("media", arred(media)),
        ("cv", arred(desvio / media if media else 0)),
        ("n_mudancas", mudancas),
    ])


def ler_procedencia(planilha):
    """Data → (boletim de origem, realinhada?), lido da aba Série longa."""
    procedencia = {}
    for linha in planilha["Série longa"].iter_rows(min_row=2, values_only=True):
        _, _, data, _, ajuste, boletim = linha
        procedencia[data.date().isoformat()] = (boletim, ajuste not in (None, "—"))
    return procedencia


def main():
    planilha = openpyxl.load_workbook(PLANILHA, read_only=True, data_only=True)
    linhas = list(planilha["Série larga"].iter_rows(values_only=True))
    procedencia = ler_procedencia(planilha)

    datas = sorted(procedencia)
    if len(datas) != len(linhas[0]) - 2:
        raise SystemExit("as duas abas discordam no número de datas")

    boletins = OrderedDict()
    for data in datas:
        boletins.setdefault(procedencia[data][0], []).append(data)
    realinhadas = [d for d in datas if procedencia[d][1]]

    produtos = OrderedDict()
    for linha in linhas[1:]:
        rotulo, unidade = linha[0], linha[1]
        if not rotulo or not unidade:
            continue  # rodapé de notas da planilha
        precos = list(linha[2:])
        if any(p is None for p in precos):
            raise SystemExit("série incompleta em %s" % rotulo)

        fator, itens, tipo = analisar_unidade(unidade)
        bloco = OrderedDict([
            ("label", ROTULOS_FIXOS.get(rotulo, rotulo)),
            ("grupo", "animal" if rotulo in PRODUTOS_ANIMAIS else "vegetal"),
            ("unidade_origem", unidade.strip()),
            ("tipo_unidade", tipo),
            ("fator_conversao_kg", arred(fator) if fator else None),
            ("itens_por_unidade", itens),
            ("precos_unidade_origem", [arred(p) for p in precos]),
        ])
        if fator:
            em_kg = [arred(p / fator) for p in precos]
            bloco["precos_rs_kg"] = em_kg
            bloco["medias_mensais_rs_kg"] = medias_mensais(datas, em_kg)
            bloco["estatisticas_rs_kg"] = estatisticas(em_kg)
        else:
            bloco["precos_rs_kg"] = None
            bloco["medias_mensais_unidade_origem"] = medias_mensais(datas, precos)
            bloco["estatisticas_unidade_origem"] = estatisticas(precos)
        produtos[slug(rotulo)] = bloco

    documento = OrderedDict([
        ("meta", OrderedDict([
            ("fonte", "EMDAGRO, Empresa de Desenvolvimento Agropecuário de Sergipe / ASPLAN"),
            ("serie", "Preços médios de atacado, CEASA Aracaju"),
            ("janela", "2026-01 a 2026-07"),
            ("geracao", "2026-08-28"),
            ("frequencia", "boletim semanal com 2 a 3 coletas por semana; não é série diária"),
            ("n_produtos", len(produtos)),
            ("n_datas", len(datas)),
            ("datas", datas),
            ("unidades", "Preço na unidade de origem do boletim. A conversão para R$/kg usa o "
                         "peso nominal da embalagem. Produtos cotados por cento ou por caixa de "
                         "dúzias ficam sem conversão: a fonte não publica peso por peça."),
            ("feriados_sem_coleta", FERIADOS),
            ("lacunas", ["Não há boletim publicado entre 16/04 e 05/05 de 2026."]),
            ("datas_realinhadas", realinhadas),
            ("nota_realinhamento",
             "Os boletins de 11/16/18 e de 23/25/30 de junho saíram com a coluna de nomes de "
             "produto reordenada sem que as colunas numéricas acompanhassem, e cada linha "
             "impressa exibia o preço de outro produto. Os valores foram realinhados para a "
             "ordem padrão e conferidos contra o boletim íntegro de 2/9/11 de junho na data "
             "coberta pelos dois (11/06): coincidem em 54 de 54 produtos. O realinhamento "
             "continua sendo inferência; a confirmação definitiva depende da ASPLAN. A "
             "transcrição literal do que foi publicado está na aba \"Junho como publicado\" da "
             "planilha de origem."),
            ("planilha_origem", "dados/fontes/precos-atacado-ceasa-se-consolidado-v3.xlsx"),
            ("boletins", boletins),
        ])),
        ("produtos", produtos),
    ])

    SAIDA.write_text(json.dumps(documento, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print("%s: %d produtos, %d datas" % (SAIDA.relative_to(RAIZ), len(produtos), len(datas)))


if __name__ == "__main__":
    main()
