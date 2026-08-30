#!/usr/bin/env python3
"""Gera dados/series-precos-se-2026.json, a série que a página /produtos/ lê.

Junta três fontes num arquivo só:

  atacado CEASA-SE
      dados/fontes/precos-atacado-ceasa-se-consolidado-v3.xlsx
      54 produtos, 48 coletas de jan a jul de 2026, transcrição dos boletins
      "Preços médios de atacado - CEASA" da EMDAGRO/ASPLAN.

  varejo Mercado Central
      dados/fontes/emdagro-mercado-central-precos-diarios-2026.csv
      65 produtos, 58 coletas de jan a jul de 2026, transcrição dos boletins
      "Preços de varejo - Mercado Central de Aracaju" (Mercado Maria Virgínia
      Leite Franco), formato longo, uma linha por produto e dia.

  varejo Augusto Franco
      dados/precos-emdagro-q1-2026.json
      11 produtos, 23 coletas de jan a mar de 2026. Único canal ainda sem
      reconsolidação: cobre só o primeiro trimestre e a cesta antiga.

Cada canal guarda a própria lista de datas, porque as coletas não coincidem e as
janelas diferem. O script não corrige, não interpola e não estima: o que não
está na fonte não é escrito.

Dependência: openpyxl (pip install openpyxl). Só é necessária para reimportar a
planilha do atacado; o site consome apenas o JSON gerado.

Uso:
    python3 scripts/importar-series-precos.py
"""

import csv
import json
import re
import unicodedata
from collections import OrderedDict
from pathlib import Path

import openpyxl

RAIZ = Path(__file__).resolve().parent.parent
PLANILHA = RAIZ / "dados" / "fontes" / "precos-atacado-ceasa-se-consolidado-v3.xlsx"
MERCADO_CENTRAL = RAIZ / "dados" / "fontes" / "emdagro-mercado-central-precos-diarios-2026.csv"
AUGUSTO_FRANCO = RAIZ / "dados" / "precos-emdagro-q1-2026.json"
SAIDA = RAIZ / "dados" / "series-precos-se-2026.json"

GERACAO = "2026-08-30"

CANAL_ATACADO = "atacado_ceasa"
CANAL_MC = "varejo_mercado_central"
CANAL_AF = "varejo_augusto_franco"

# Datas que aparecem no cabeçalho dos boletins de atacado marcadas como feriado,
# sem preço publicado. Não entram na série.
FERIADOS = ["2026-04-02", "2026-04-21", "2026-06-04"]

# Os sete últimos produtos do boletim de atacado formam o bloco de origem animal.
PRODUTOS_ANIMAIS = {
    "Carne bovina dianteira",
    "Carne bovina traseira",
    "Frango de granja",
    "Manteiga",
    "Ovo branco grande",
    "Ovo vermelho grande",
    "Queijo coalho",
}

# Slugs que já circulam no portal (mural da capa, grade dos panoramas, links de
# produto) e não podem mudar, mais os casos em que a transliteração automática
# ficaria ilegível.
SLUGS_FIXOS = {
    "Macaxeira (Aipim)": "macaxeira",
    "Amendoim c/ casca cozido": "amendoim_com_casca_cozido",
}

# Rótulos já publicados no portal, mantidos para não divergir.
ROTULOS_FIXOS = {
    "Macaxeira (Aipim)": "Macaxeira (aipim)",
}

# O boletim de varejo grafa tudo em caixa alta e sem acento. Estes são os
# rótulos legíveis dos produtos que só existem no varejo — ortografia, não
# renomeação: nenhum produto é fundido com outro por aqui.
ROTULOS_VAREJO = {
    "ABACAXI GRANDE": "Abacaxi grande",
    "ABOBORA": "Abóbora",
    "ACEM C/ OSSO": "Acém com osso",
    "ACEM S/ OSSO": "Acém sem osso",
    "ALCATRA": "Alcatra",
    "ALFACE": "Alface",
    "CARNE SUINA TRASEIRA": "Carne suína traseira",
    "CEBOLA BRANCA": "Cebola branca",
    "CHA DE DENTRO (COXAO MOLE)": "Chã de dentro (coxão mole)",
    "CHA DE FORA (COXAO DURO)": "Chã de fora (coxão duro)",
    "CONTRA FILE": "Contrafilé",
    "COSTELA": "Costela",
    "FILE MIGNION": "Filé mignon",
    "MUSCULO": "Músculo",
    "PALETA C/ OSSO": "Paleta com osso",
    "PATINHO": "Patinho",
    "PEITO C/ OSSO": "Peito com osso",
}

# Pareamentos entre os dois boletins que são só diferença de grafia. Casos em
# que o varejo usa nome genérico e o atacado especifica variedade — abacaxi
# grande, abóbora, alface, cebola branca — ficam de fora de propósito: uni-los
# afirmaria uma identidade de produto que a fonte não declara.
PAREAMENTOS = {
    "ARROZ AGULHA T 1": "arroz_agulha_t1",
    "MAMAO HAWAY": "mamao_hawai",
}

ROTULOS_CANAL = {
    CANAL_ATACADO: "Atacado CEASA-SE",
    CANAL_MC: "Varejo Mercado Central",
    CANAL_AF: "Varejo Augusto Franco",
}

# Nome curto para onde a largura é apertada, como o cabeçalho dos cartões.
ROTULOS_CURTOS = {
    CANAL_ATACADO: "CEASA-SE",
    CANAL_MC: "Mercado Central",
    CANAL_AF: "Augusto Franco",
}

# Unidades de contagem: o boletim cota por peça, não por peso. O valor é
# (peças por unidade, nome da peça). Cento e caixa de dúzias são embalagens de
# peças, e dividir por 100 ou por 360 é aritmética da própria unidade — não
# estimativa. O que essa divisão assume é que o cento conta a mesma peça que o
# varejo vende: 100 molhos de coentro, 100 pés de alface. A nota metodológica
# declara essa premissa.
UNIDADES_CONTAGEM = {
    "cento": (100, "unidade"),
    "cx 30 dz": (360, "unidade"),
    "dz": (12, "unidade"),
    "um": (1, "unidade"),
    "uma": (1, "unidade"),
    "unidade": (1, "unidade"),
    "pe": (1, "pé"),
    "cab.": (1, "cabeça"),
    "molho": (1, "molho"),
    "espiga": (1, "espiga"),
}

# A lata não é peça nem peso: é recipiente de conteúdo não declarado. Não
# converte para nada, e o preço fica só na unidade de origem.
UNIDADES_OPACAS = {"lata"}

# Quando dois canais nomeiam a peça de formas diferentes, vale a mais específica:
# "R$ por molho" diz mais ao leitor do que "R$ por unidade".
PECAS_ESPECIFICAS = ("molho", "pé", "espiga", "cabeça")


def slug(rotulo):
    if rotulo in SLUGS_FIXOS:
        return SLUGS_FIXOS[rotulo]
    return re.sub(r"_+", "_", re.sub(r"[^a-z0-9]+", "_", sem_acento(rotulo).lower())).strip("_")


def sem_acento(texto):
    return unicodedata.normalize("NFKD", str(texto)).encode("ascii", "ignore").decode()


def normalizar_nome(texto):
    return re.sub(r"[^a-z0-9]+", " ", sem_acento(texto).lower()).strip()


def analisar_unidade(unidade):
    """Devolve (fator_kg, itens_por_unidade, tipo, peca) para a unidade do boletim.

    fator_kg é None quando a unidade é de contagem: nesse caso o preço fica na
    unidade de origem e a página precisa dizer isso ao leitor.
    """
    chave = unidade.strip().lower()
    if chave in UNIDADES_OPACAS:
        return None, None, "opaca", None
    if chave in UNIDADES_CONTAGEM:
        itens, peca = UNIDADES_CONTAGEM[chave]
        return None, itens, "unidade", peca
    achado = re.search(r"(\d+(?:[.,]\d+)?)\s*kg", chave)
    if achado:
        return float(achado.group(1).replace(",", ".")), None, "peso", None
    if chave == "kg":
        return 1.0, None, "peso", None
    raise ValueError("unidade não reconhecida: %r" % unidade)


def arred(valor, casas=4):
    arredondado = round(valor, casas)
    return int(arredondado) if arredondado == int(arredondado) else arredondado


def medias_mensais(datas, precos):
    acumulado = OrderedDict()
    for data, preco in zip(datas, precos):
        if preco is None:
            continue
        acumulado.setdefault(data[:7], []).append(preco)
    return OrderedDict((mes, arred(sum(v) / len(v))) for mes, v in acumulado.items())


def estatisticas(precos):
    serie = [p for p in precos if p is not None]
    media = sum(serie) / len(serie)
    # Desvio-padrão amostral, como na aba de estatísticas da planilha de origem.
    desvio = (sum((p - media) ** 2 for p in serie) / (len(serie) - 1)) ** 0.5 if len(serie) > 1 else 0
    return OrderedDict([
        ("n_observacoes", len(serie)),
        ("minimo", arred(min(serie))),
        ("maximo", arred(max(serie))),
        ("media", arred(media)),
        ("cv", arred(desvio / media if media else 0)),
        ("n_mudancas", sum(1 for a, b in zip(serie, serie[1:]) if a != b)),
    ])


def bloco_de_serie(datas, unidades, precos):
    """Descreve a série de um canal na unidade em que o boletim a publica.

    A conversão para a unidade de referência do produto acontece depois, em
    converter_para_referencia(): ela depende dos outros canais, que este bloco
    ainda não conhece.
    """
    distintas = []
    for u in unidades:
        if not distintas or distintas[-1] != u:
            distintas.append(u)
    variou = len(set(unidades)) > 1

    bloco = OrderedDict()
    bloco["unidade_origem"] = distintas[0] if not variou else " → ".join(distintas)
    bloco["unidade_variou"] = variou
    if variou:
        bloco["unidades_por_data"] = list(unidades)
    bloco["precos_unidade_origem"] = [arred(p) for p in precos]

    if not variou:
        fator, itens, tipo, peca = analisar_unidade(unidades[0])
        bloco["tipo_unidade"] = tipo
        bloco["peca_contada"] = peca
        bloco["fator_conversao_kg"] = arred(fator) if fator else None
        bloco["itens_por_unidade"] = itens
    else:
        bloco["tipo_unidade"] = "mista"
        bloco["peca_contada"] = None
        bloco["fator_conversao_kg"] = None
        bloco["itens_por_unidade"] = None

    bloco["_datas"] = datas
    bloco["_unidades"] = list(unidades)
    return bloco


def unidade_de_referencia(ficha, canais):
    """Escolhe a unidade única em que o produto será publicado e comparado.

    Peso ganha de contagem: R$/kg é o padrão do portal e é o que a maior parte
    da série suporta. Só quando nenhum canal cota por peso — coco, coentro, ovo —
    a referência passa a ser o preço por peça.
    """
    tem_peso = False
    peca = None
    for canal in canais:
        bloco = ficha.get(canal)
        if not bloco:
            continue
        for unidade in bloco["_unidades"]:
            fator, _, tipo, nome = analisar_unidade(unidade)
            if fator:
                tem_peso = True
            elif nome and (peca is None or (nome in PECAS_ESPECIFICAS and peca == "unidade")):
                peca = nome

    if tem_peso:
        return OrderedDict([("tipo", "kg"), ("peca", None),
                            ("sufixo", "/kg"), ("legenda", "R$ por quilo")])
    peca = peca or "unidade"
    return OrderedDict([("tipo", "peca"), ("peca", peca),
                        ("sufixo", "/" + peca), ("legenda", "R$ por " + peca)])


def converter_para_referencia(bloco, referencia):
    """Reescreve a série do canal na unidade de referência do produto.

    O preço de uma data vira null quando a unidade daquela data não converte:
    quilo não vira peça sem peso por peça, e peça não vira quilo pelo mesmo
    motivo. Preferimos o buraco à suposição — e é o buraco que faz a linha do
    gráfico se interromper onde a comparação deixaria de valer.
    """
    valores = []
    conversoes = []
    for unidade, preco in zip(bloco["_unidades"], bloco["precos_unidade_origem"]):
        fator, itens, tipo, _ = analisar_unidade(unidade)
        if referencia["tipo"] == "kg":
            valores.append(arred(preco / fator) if fator else None)
            conversoes.append("÷ %g kg" % fator if fator else None)
        else:
            valores.append(arred(preco / itens) if itens else None)
            conversoes.append(("÷ %d peças" % itens) if itens and itens > 1
                              else ("igual" if itens else None))

    bloco["precos_referencia"] = valores
    bloco["conversoes"] = sorted({c for c in conversoes if c})
    bloco["convertivel"] = any(v is not None for v in valores)

    # Trechos contínuos de valor convertido: cada um vira uma linha no gráfico,
    # e o corte entre eles marca onde a série muda de unidade.
    trechos = []
    inicio = None
    for i, valor in enumerate(valores + [None]):
        if valor is not None and inicio is None:
            inicio = i
        elif valor is None and inicio is not None:
            trechos.append(OrderedDict([
                ("inicio", bloco["_datas"][inicio]),
                ("fim", bloco["_datas"][i - 1]),
                ("indice_inicio", inicio),
                ("indice_fim", i - 1),
                ("unidade", bloco["_unidades"][inicio]),
                ("estatisticas", estatisticas(valores[inicio:i])),
            ]))
            inicio = None
    bloco["trechos"] = trechos

    if bloco["convertivel"]:
        limpos = [v for v in valores if v is not None]
        datas_limpas = [d for d, v in zip(bloco["_datas"], valores) if v is not None]
        bloco["medias_mensais"] = medias_mensais(datas_limpas, limpos)
        bloco["estatisticas"] = estatisticas(limpos)
        bloco["n_convertidas"] = len(limpos)
    else:
        bloco["medias_mensais"] = None
        bloco["estatisticas"] = None
        bloco["n_convertidas"] = 0

    del bloco["_datas"]
    del bloco["_unidades"]
    return bloco


def ler_procedencia(planilha):
    """Data → (boletim de origem, realinhada?), lido da aba Série longa."""
    procedencia = {}
    for linha in planilha["Série longa"].iter_rows(min_row=2, values_only=True):
        _, _, data, _, ajuste, boletim = linha
        procedencia[data.date().isoformat()] = (boletim, ajuste not in (None, "—"))
    return procedencia


def ler_atacado():
    """Devolve (datas, boletins, datas_realinhadas, {slug: (rotulo, bloco)})."""
    planilha = openpyxl.load_workbook(PLANILHA, read_only=True, data_only=True)
    linhas = list(planilha["Série larga"].iter_rows(values_only=True))
    procedencia = ler_procedencia(planilha)

    datas = sorted(procedencia)
    if len(datas) != len(linhas[0]) - 2:
        raise SystemExit("as duas abas da planilha discordam no número de datas")

    boletins = OrderedDict()
    for data in datas:
        boletins.setdefault(procedencia[data][0], []).append(data)

    series = OrderedDict()
    for linha in linhas[1:]:
        rotulo, unidade = linha[0], linha[1]
        if not rotulo or not unidade:
            continue  # rodapé de notas da planilha
        precos = list(linha[2:])
        if any(p is None for p in precos):
            raise SystemExit("série de atacado incompleta em %s" % rotulo)
        bloco = bloco_de_serie(datas, [unidade.strip()] * len(datas), precos)
        series[slug(rotulo)] = (rotulo, bloco)

    return datas, boletins, [d for d in datas if procedencia[d][1]], series


def ler_mercado_central(rotulos_atacado):
    """Lê o CSV longo do Mercado Central.

    Devolve (datas, boletins, {slug: (rotulo, bloco)}). Duas datas aparecem em
    dois boletins cada, porque as janelas publicadas se sobrepõem; os valores
    coincidem, e a leitura repetida é descartada em vez de virar coleta a mais.
    """
    with MERCADO_CENTRAL.open(encoding="utf-8") as arquivo:
        linhas = list(csv.DictReader(arquivo))

    observacoes = {}
    boletins = OrderedDict()
    for linha in linhas:
        data, produto = linha["data"], linha["produto"]
        registro = (linha["unidade"].strip(), float(linha["preco"]))
        anterior = observacoes.get((data, produto))
        if anterior and anterior != registro:
            raise SystemExit("leituras conflitantes em %s, %s: %r vs %r"
                             % (produto, data, anterior, registro))
        observacoes[(data, produto)] = registro
        boletins.setdefault(linha["arquivo"], set()).add(data)

    datas = sorted({d for d, _ in observacoes})
    produtos = sorted({p for _, p in observacoes})

    series = OrderedDict()
    for produto in produtos:
        faltando = [d for d in datas if (d, produto) not in observacoes]
        if faltando:
            raise SystemExit("%s sem preço em %d datas" % (produto, len(faltando)))
        unidades = [observacoes[(d, produto)][0] for d in datas]
        precos = [observacoes[(d, produto)][1] for d in datas]
        series[chave_de_varejo(produto, rotulos_atacado)] = (produto, bloco_de_serie(datas, unidades, precos))

    boletins = OrderedDict((arquivo, sorted(datas))
                           for arquivo, datas in sorted(boletins.items(), key=lambda kv: min(kv[1])))
    return datas, boletins, series


def chave_de_varejo(produto, rotulos_atacado):
    """Slug do produto de varejo: o mesmo do atacado quando é o mesmo produto."""
    if produto in PAREAMENTOS:
        return PAREAMENTOS[produto]
    normalizado = normalizar_nome(produto)
    if normalizado in rotulos_atacado:
        return rotulos_atacado[normalizado]
    return slug(ROTULOS_VAREJO.get(produto, produto.title()))


def main():
    datas_atacado, boletins_atacado, realinhadas, series_atacado = ler_atacado()

    produtos = OrderedDict()
    rotulos_atacado = {}
    for chave, (rotulo, bloco) in series_atacado.items():
        produtos[chave] = OrderedDict([
            ("label", ROTULOS_FIXOS.get(rotulo, rotulo)),
            ("grupo", "animal" if rotulo in PRODUTOS_ANIMAIS else "vegetal"),
            (CANAL_ATACADO, bloco),
        ])
        rotulos_atacado[normalizar_nome(produtos[chave]["label"])] = chave

    datas_mc, boletins_mc, series_mc = ler_mercado_central(rotulos_atacado)
    animais_varejo = set()
    with MERCADO_CENTRAL.open(encoding="utf-8") as arquivo:
        for linha in csv.DictReader(arquivo):
            if linha["grupo"] == "Origem animal":
                animais_varejo.add(linha["produto"])

    for chave, (rotulo, bloco) in series_mc.items():
        if chave not in produtos:
            produtos[chave] = OrderedDict([
                ("label", ROTULOS_VAREJO.get(rotulo, rotulo.title())),
                ("grupo", "animal" if rotulo in animais_varejo else "vegetal"),
            ])
        produtos[chave][CANAL_MC] = bloco

    # Augusto Franco continua vindo da consolidação do primeiro trimestre: é o
    # único canal ainda sem reconsolidação a partir dos boletins.
    augusto = json.loads(AUGUSTO_FRANCO.read_text(encoding="utf-8"))
    datas_af = augusto["meta"]["canais"][CANAL_AF]["datas"]
    for chave, ficha in augusto["produtos"].items():
        origem = ficha.get(CANAL_AF)
        if not origem:
            continue
        if chave not in produtos:
            raise SystemExit("produto de varejo sem correspondência: %s" % chave)
        fator = origem["fator_conversao_kg"]
        precos = [p * fator for p in origem["precos_rs_kg"]]
        unidade = origem["unidade_origem"].strip()
        produtos[chave][CANAL_AF] = bloco_de_serie(datas_af, [unidade] * len(datas_af), precos)

    # Com todos os canais de cada produto na mão, escolhe a unidade única em que
    # ele será publicado e reescreve cada canal nela.
    canais_todos = (CANAL_ATACADO, CANAL_MC, CANAL_AF)
    for ficha in produtos.values():
        referencia = unidade_de_referencia(ficha, canais_todos)
        ficha["unidade_referencia"] = referencia
        for canal in canais_todos:
            if canal in ficha:
                converter_para_referencia(ficha[canal], referencia)

    produtos = OrderedDict(sorted(produtos.items(), key=lambda kv: normalizar_nome(kv[1]["label"])))

    canais = OrderedDict()
    canais[CANAL_ATACADO] = OrderedDict([
        ("label", ROTULOS_CANAL[CANAL_ATACADO]),
        ("label_curto", ROTULOS_CURTOS[CANAL_ATACADO]),
        ("descricao", "Atacado CEASA-SE, Centrais de Abastecimento de Sergipe (Aracaju)"),
        ("frequencia", "boletim semanal com 2 a 3 coletas por semana; não é série diária"),
        ("janela", "2026-01 a 2026-07"),
        ("n_datas", len(datas_atacado)),
        ("n_produtos", len(series_atacado)),
        ("datas", datas_atacado),
        ("nota", "Preços originais por embalagem (caixa, saco, arroba) convertidos para R$/kg "
                 "pelo peso nominal da embalagem."),
        ("feriados_sem_coleta", FERIADOS),
        ("lacunas", ["Não há boletim publicado entre 16/04 e 05/05 de 2026."]),
        ("datas_realinhadas", realinhadas),
        ("nota_realinhamento",
         "Os boletins de 11/16/18 e de 23/25/30 de junho saíram com a coluna de nomes de produto "
         "reordenada sem que as colunas numéricas acompanhassem, e cada linha impressa exibia o "
         "preço de outro produto. Os valores foram realinhados para a ordem padrão e conferidos "
         "contra o boletim íntegro de 2/9/11 de junho na data coberta pelos dois (11/06): "
         "coincidem em 54 de 54 produtos. O realinhamento é inferência verificada, não "
         "confirmação da fonte."),
        ("boletins", boletins_atacado),
    ])
    canais[CANAL_MC] = OrderedDict([
        ("label", ROTULOS_CANAL[CANAL_MC]),
        ("label_curto", ROTULOS_CURTOS[CANAL_MC]),
        ("descricao", "Varejo Mercado Maria Virgínia Leite Franco (Mercado Central de Aracaju)"),
        ("frequencia", "boletim semanal com 2 a 3 coletas por semana; não é série diária"),
        ("janela", "2026-01 a 2026-07"),
        ("n_datas", len(datas_mc)),
        ("n_produtos", len(series_mc)),
        ("datas", datas_mc),
        ("nota", "Mercado popular com preços próximos do atacado: canal híbrido, não proxy de "
                 "supermercado. Em sete produtos a unidade de venda muda ao longo do período, e "
                 "esses trechos não se comparam entre si."),
        ("boletins", boletins_mc),
    ])
    origem_af = augusto["meta"]["canais"][CANAL_AF]
    canais[CANAL_AF] = OrderedDict([
        ("label", ROTULOS_CANAL[CANAL_AF]),
        ("label_curto", ROTULOS_CURTOS[CANAL_AF]),
        ("descricao", origem_af["descricao"]),
        ("frequencia", origem_af["frequencia"]),
        ("janela", "2026-01 a 2026-03"),
        ("n_datas", origem_af["n_datas"]),
        ("n_produtos", sum(1 for p in produtos.values() if CANAL_AF in p)),
        ("datas", datas_af),
        ("nota", origem_af.get("nota", "")),
    ])

    documento = OrderedDict([
        ("meta", OrderedDict([
            ("fonte", "EMDAGRO, Empresa de Desenvolvimento Agropecuário de Sergipe / ASPLAN"),
            ("serie", "Preços médios de atacado e varejo, Aracaju"),
            ("janela", "2026-01 a 2026-07"),
            ("geracao", GERACAO),
            ("unidade_padrao", "R$/kg"),
            ("n_produtos", len(produtos)),
            ("unidades", "Preço na unidade de origem do boletim. A conversão para R$/kg usa o "
                         "peso nominal da embalagem. Produtos cotados por peça — cento, dúzia, "
                         "molho, pé, cabeça — ficam sem conversão: a fonte não publica peso por "
                         "peça. Cada produto tem uma unidade de referência, em unidade_referencia, "
                         "e todos os seus canais são publicados nela: onde a conversão não existe, "
                         "o preço fica nulo em vez de ser suposto."),
            ("canais", canais),
            ("arquivos_de_origem", OrderedDict([
                (CANAL_ATACADO, "dados/fontes/precos-atacado-ceasa-se-consolidado-v3.xlsx"),
                (CANAL_MC, "dados/fontes/emdagro-mercado-central-precos-diarios-2026.csv"),
                (CANAL_AF, "dados/precos-emdagro-q1-2026.json"),
            ])),
        ])),
        ("produtos", produtos),
    ])

    SAIDA.write_text(json.dumps(documento, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")

    def conta(canal):
        return sum(1 for p in produtos.values() if canal in p)

    print("%s: %d produtos" % (SAIDA.relative_to(RAIZ), len(produtos)))
    referencias = {}
    for ficha in produtos.values():
        referencias[ficha["unidade_referencia"]["legenda"]] = referencias.get(
            ficha["unidade_referencia"]["legenda"], 0) + 1
    print("  unidades de referência: %s" % ", ".join(
        "%s (%d)" % (k, v) for k, v in sorted(referencias.items())))
    parciais = [(p["label"], c) for p in produtos.values() for c in canais_todos
                if c in p and p[c]["n_convertidas"] < len(p[c]["precos_unidade_origem"])]
    print("  canais com trecho fora da unidade de referência: %d" % len(parciais))
    for rotulo, canal in parciais:
        ficha = [p for p in produtos.values() if p["label"] == rotulo][0]
        bloco = ficha[canal]
        print("     %-22s %-24s %d de %d coletas · %s" % (
            rotulo, canal, bloco["n_convertidas"], len(bloco["precos_unidade_origem"]),
            bloco["unidade_origem"]))
    print("  atacado CEASA-SE      %2d produtos · %d coletas" % (conta(CANAL_ATACADO), len(datas_atacado)))
    print("  varejo Mercado Central %2d produtos · %d coletas" % (conta(CANAL_MC), len(datas_mc)))
    print("  varejo Augusto Franco  %2d produtos · %d coletas" % (conta(CANAL_AF), len(datas_af)))
    variaveis = [p["label"] for p in produtos.values()
                 if any(c in p and p[c].get("unidade_variou") for c in (CANAL_ATACADO, CANAL_MC, CANAL_AF))]
    print("  unidade variável em %d produtos: %s" % (len(variaveis), ", ".join(variaveis)))


if __name__ == "__main__":
    main()
